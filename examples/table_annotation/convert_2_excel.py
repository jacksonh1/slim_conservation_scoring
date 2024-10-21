import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Step 1: Read the CSV file into a pandas DataFrame
csv_file = "./table_ANNOTATED.csv"
df = pd.read_csv(csv_file)

# Step 2: List of columns that contain file paths
file_path_columns = [
    "pairk_aln_needleman_lf5_rf5_edssmat50-Vertebrata-aln_slice_file",
    "pairk_aln_needleman_lf5_rf5_edssmat50-multi_level_plot",
    "aln_property_entropy-Vertebrata-aln_slice_file",
    "aln_property_entropy-multi_level_plot",
]

# Step 3: Create an Excel writer object and write the DataFrame to Excel
excel_file = "output_file.xlsx"
df.to_excel(excel_file, index=False)

# Step 4: Open the Excel file with openpyxl to add hyperlinks
wb = load_workbook(excel_file)
ws = wb.active

# Step 5: Loop through each specified column and each row to add clickable file links
for col_name in file_path_columns:
    col_idx = (
        df.columns.get_loc(col_name) + 1
    )  # Get the column index (1-based for openpyxl)

    for row_idx, file_path in enumerate(
        df[col_name], start=2
    ):  # Start at row 2 because row 1 is the header
        if pd.notna(file_path):  # Only add hyperlink if the file path is not NaN
            # Construct the absolute file path
            # absolute_path = os.path.abspath(file_path)
            print(file_path)
            link = f"{file_path}"  # Create a file link
            print(link)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.hyperlink = link  # Add hyperlink to the cell
            cell.value = file_path  # Ensure the cell displays the file path
            cell.font = Font(
                color="0000FF", underline="single"
            )  # Format as a clickable link

# Step 6: Save the modified Excel file
wb.save(excel_file)
