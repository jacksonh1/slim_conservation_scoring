import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path
import os


def convert_csv_to_excel(csv_file: str | Path):
    csv_file = Path(csv_file)
    excel_file = csv_file.with_suffix(".xlsx")
    df = pd.read_csv(csv_file)
    file_path_columns = [
        col
        for col in df.columns
        if "multi_level_plot" in col or "aln_slice_file" in col
    ]
    df.to_excel(excel_file, index=False)
    wb = load_workbook(excel_file)
    ws = wb.active
    for col_name in file_path_columns:
        col_idx = (
            df.columns.get_loc(col_name) + 1
        )  # Get the column index (1-based for openpyxl)
        for row_idx, file_path in enumerate(
            df[col_name], start=2
        ):  # Start at row 2 because row 1 is the header
            if pd.notna(file_path):  # Only add hyperlink if the file path is not NaN
                # Construct the absolute file path
                absolute_path = os.path.abspath(file_path)
                link = f"file://{absolute_path}"  # Create a file link
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.hyperlink = link  # Add hyperlink to the cell
                cell.value = file_path  # Ensure the cell displays the file path
                cell.font = Font(
                    color="0000FF", underline="single"
                )  # Format as a clickable link
    wb.save(excel_file)
